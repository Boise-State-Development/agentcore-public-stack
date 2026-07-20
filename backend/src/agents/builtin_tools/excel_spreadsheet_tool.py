"""Excel spreadsheet tools (create / modify / list / read).

Each tool runs openpyxl code inside AWS Bedrock Code Interpreter and uses the
existing user-files store (``apis.shared.files``) for persistence and delivery
— generated/modified ``.xlsx`` files land in ``S3_USER_FILES_BUCKET_NAME`` with
a ``FileMetadata`` row (status READY) in ``DYNAMODB_USER_FILES_TABLE_NAME``, so
they appear in the chat's Files panel and are downloadable via the app-api
``/files/{id}/preview-url`` route.

Tools
-----
* ``create_excel_spreadsheet`` — build a new workbook from openpyxl code.
* ``modify_excel_spreadsheet`` — edit an existing workbook with openpyxl code.
* ``list_excel_spreadsheets``  — list the .xlsx files available in this chat.
* ``read_excel_spreadsheet``   — extract an existing workbook's cell values.

This is the create/modify/read/list toolset for *generated* ``.xlsx`` files. It
is distinct from the spreadsheet *analysis* tools (``list_spreadsheets`` /
``analyze_spreadsheet`` in ``builtin_tools.spreadsheet_analysis``), which read
and aggregate existing uploaded/knowledge-base tabular files with pandas.

Design notes
------------
* The Code Interpreter + user-files storage plumbing is shared with the Word
  toolset and lives in ``builtin_tools.office._storage``; this module keeps
  only the openpyxl specifics (preamble, generate/modify/extract) and the four
  tool factories.
* Identity (``user_id`` / ``session_id``) is captured by closure via the
  ``make_*`` factories — the same pattern used by the artifacts, Word document,
  and spreadsheet_analysis tools (the Strands runtime here does NOT populate
  ``ToolContext.invocation_state`` with identity). The tools are injected
  per-request through ``extra_tools`` (see ``_build_excel_spreadsheet_tools`` in
  ``apis/inference_api/chat/routes.py``); they are deliberately NOT registered
  in ``builtin_tools/__init__`` because they need request-scoped identity.
"""

from __future__ import annotations

import asyncio
import logging
from typing import Any, Dict, Optional

from strands import tool

from agents.builtin_tools.office._storage import (
    _DocGenError,
    _ci_exec,
    _ci_read_bytes,
    _ci_write_bytes,
    _download_card,
    _download_s3_bytes,
    _error,
    _get_code_interpreter_id,
    _NO_CI_MESSAGE,
    _region,
    _storage_configured,
    _store_document,
    _validate_document_name,
)

logger = logging.getLogger(__name__)

# Excel workbook MIME type (matches apis.shared.files.ALLOWED_MIME_TYPES).
_XLSX_MIME = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

# Sandbox path used to stage a source workbook loaded from S3.
_SANDBOX_SOURCE = "_source.xlsx"

_NO_STORAGE_MESSAGE = (
    "❌ Excel spreadsheet storage is not configured "
    "(S3_USER_FILES_BUCKET_NAME is not set on the runtime)."
)


# ---------------------------------------------------------------------------
# openpyxl workbook builders (run in Code Interpreter)
# ---------------------------------------------------------------------------


_XLSX_PREAMBLE = (
    "from openpyxl import Workbook, load_workbook\n"
    "from openpyxl.styles import Font, PatternFill, Alignment, Border, Side\n"
    "from openpyxl.utils import get_column_letter\n"
    "from openpyxl.chart import BarChart, LineChart, PieChart, Reference\n"
)


def _generate_xlsx_bytes(
    code_interpreter_id: str, python_code: str, filename: str
) -> bytes:
    """Build a new .xlsx from user code and return its bytes.

    Blocking (boto3 / Code Interpreter) — call via ``asyncio.to_thread``.
    """
    from bedrock_agentcore.tools.code_interpreter_client import CodeInterpreter

    code_interpreter = CodeInterpreter(_region())
    code_interpreter.start(identifier=code_interpreter_id)
    try:
        # The user's code operates on a pre-initialized ``wb`` / ``ws`` and must
        # not call Workbook()/wb.save() itself — we own the lifecycle.
        _ci_exec(
            code_interpreter,
            (
                f"{_XLSX_PREAMBLE}\n"
                "wb = Workbook()\n"
                "ws = wb.active\n\n"
                f"{python_code}\n\n"
                f"wb.save({filename!r})\n"
            ),
        )
        data = _ci_read_bytes(code_interpreter, filename)
        if data is None:
            raise _DocGenError(
                f"Workbook '{filename}' was not produced. Make sure your code "
                "writes cells to `ws` (or another sheet on `wb`)."
            )
        return data
    finally:
        try:
            code_interpreter.stop()
        except Exception:  # pragma: no cover - cleanup best-effort
            pass


def _modify_xlsx_bytes(
    code_interpreter_id: str,
    source_bytes: bytes,
    python_code: str,
    output_filename: str,
) -> bytes:
    """Load an existing .xlsx, apply user edits, return the new bytes.

    Blocking — call via ``asyncio.to_thread``.
    """
    from bedrock_agentcore.tools.code_interpreter_client import CodeInterpreter

    code_interpreter = CodeInterpreter(_region())
    code_interpreter.start(identifier=code_interpreter_id)
    try:
        _ci_write_bytes(code_interpreter, _SANDBOX_SOURCE, source_bytes)
        _ci_exec(
            code_interpreter,
            (
                f"{_XLSX_PREAMBLE}\n"
                f"wb = load_workbook({_SANDBOX_SOURCE!r})\n"
                "ws = wb.active\n\n"
                f"{python_code}\n\n"
                f"wb.save({output_filename!r})\n"
            ),
        )
        data = _ci_read_bytes(code_interpreter, output_filename)
        if data is None:
            raise _DocGenError(
                f"Modified workbook '{output_filename}' was not produced."
            )
        return data
    finally:
        try:
            code_interpreter.stop()
        except Exception:  # pragma: no cover - cleanup best-effort
            pass


def _extract_xlsx_text(code_interpreter_id: str, source_bytes: bytes) -> str:
    """Extract readable cell values (per sheet, pipe-delimited rows) from a .xlsx.

    Uses ``data_only=True`` so cached formula results are shown when present.
    Blocking — call via ``asyncio.to_thread``.
    """
    from bedrock_agentcore.tools.code_interpreter_client import CodeInterpreter

    code_interpreter = CodeInterpreter(_region())
    code_interpreter.start(identifier=code_interpreter_id)
    try:
        _ci_write_bytes(code_interpreter, _SANDBOX_SOURCE, source_bytes)
        extraction = (
            "from openpyxl import load_workbook\n"
            f"wb = load_workbook({_SANDBOX_SOURCE!r}, data_only=True)\n"
            "lines = []\n"
            "for ws in wb.worksheets:\n"
            "    lines.append('## Sheet: ' + str(ws.title))\n"
            "    empty = True\n"
            "    for row in ws.iter_rows(values_only=True):\n"
            "        cells = ['' if c is None else str(c) for c in row]\n"
            "        while cells and cells[-1] == '':\n"
            "            cells.pop()\n"
            "        if not cells:\n"
            "            continue\n"
            "        empty = False\n"
            "        lines.append(' | '.join(cells))\n"
            "    if empty:\n"
            "        lines.append('(empty sheet)')\n"
            "    lines.append('')\n"
            "print('\\n'.join(lines))\n"
        )
        return _ci_exec(code_interpreter, extraction).strip()
    finally:
        try:
            code_interpreter.stop()
        except Exception:  # pragma: no cover - cleanup best-effort
            pass


# ---------------------------------------------------------------------------
# User-files lookup
# ---------------------------------------------------------------------------


async def _find_excel_spreadsheet(
    user_id: str, session_id: str, spreadsheet_name: str
):
    """Find the newest READY .xlsx in this session matching ``spreadsheet_name``.

    Returns the ``FileMetadata`` or ``None``. ``list_session_files`` returns
    newest-first, so the first match is the latest version.
    """
    from apis.shared.files import FileStatus, get_file_upload_repository

    target = (
        spreadsheet_name
        if spreadsheet_name.lower().endswith(".xlsx")
        else f"{spreadsheet_name}.xlsx"
    )
    files = await get_file_upload_repository().list_session_files(
        session_id, status=FileStatus.READY
    )
    for meta in files:
        if (
            meta.user_id == user_id
            and meta.mime_type == _XLSX_MIME
            and meta.filename.lower() == target.lower()
        ):
            return meta
    return None


# ---------------------------------------------------------------------------
# Tool factories
# ---------------------------------------------------------------------------


def make_create_excel_spreadsheet_tool(session_id: str, user_id: str):
    """Create a ``create_excel_spreadsheet`` tool bound to the given identity."""

    @tool
    async def create_excel_spreadsheet(
        python_code: str,
        spreadsheet_name: str,
    ) -> Any:
        """Create a new Excel (.xlsx) spreadsheet using openpyxl code.

        Executes openpyxl code in a sandboxed Code Interpreter to build a
        workbook from scratch, saves it to the user's files, and returns a
        download card. Great for tabular data, multi-sheet workbooks, formatted
        headers, formulas, and native Excel charts.

        Available libraries in the sandbox: openpyxl, pandas, numpy.

        Args:
            python_code: openpyxl code that builds the workbook. A blank
                workbook is already available as ``wb = Workbook()`` with its
                first sheet as ``ws = wb.active`` — do NOT call ``Workbook()``
                or ``wb.save()`` yourself; the tool saves it for you.
                ``Workbook``, ``load_workbook``, ``Font``, ``PatternFill``,
                ``Alignment``, ``Border``, ``Side``, ``get_column_letter`` and
                the chart classes (``BarChart``, ``LineChart``, ``PieChart``,
                ``Reference``) are already imported.

                Example (headers + rows + a formula):
                    ws.title = 'Sales'
                    ws.append(['Quarter', 'Revenue'])
                    ws['A1'].font = Font(bold=True)
                    ws['B1'].font = Font(bold=True)
                    ws.append(['Q1', 100])
                    ws.append(['Q2', 120])
                    ws['B4'] = '=SUM(B2:B3)'

                Example (add a second sheet + a bar chart):
                    ws2 = wb.create_sheet('Chart')
                    data = Reference(ws, min_col=2, min_row=1, max_row=3)
                    cats = Reference(ws, min_col=1, min_row=2, max_row=3)
                    chart = BarChart()
                    chart.add_data(data, titles_from_data=True)
                    chart.set_categories(cats)
                    ws2.add_chart(chart, 'A1')

                To load data from a pandas DataFrame:
                    from openpyxl.utils.dataframe import dataframe_to_rows
                    for r in dataframe_to_rows(df, index=False, header=True):
                        ws.append(r)

            spreadsheet_name: File name WITHOUT extension (.xlsx is added
                automatically). Use only letters, numbers, hyphens, and
                underscores (e.g. "sales-2026", "Q4_budget").

        Returns:
            An inline download card. The workbook is also saved to this
            chat's Files.
        """
        is_valid, error_msg = _validate_document_name(spreadsheet_name)
        if not is_valid:
            return _error(
                f"❌ Invalid spreadsheet name '{spreadsheet_name}': {error_msg}\n\n"
                "Examples: sales-2026, Q4_budget, report-final"
            )

        filename = f"{spreadsheet_name}.xlsx"
        code_interpreter_id = _get_code_interpreter_id()
        if not code_interpreter_id:
            return _error(_NO_CI_MESSAGE)
        if not _storage_configured():
            return _error(_NO_STORAGE_MESSAGE)

        try:
            file_bytes = await asyncio.to_thread(
                _generate_xlsx_bytes, code_interpreter_id, python_code, filename
            )
        except _DocGenError as exc:
            return _error(
                f"❌ Failed to create '{filename}'.\n\n```\n{exc}\n```\n\n"
                "Check the openpyxl code for errors."
            )
        except Exception as exc:  # noqa: BLE001 - surface any sandbox error
            logger.error(f"create_excel_spreadsheet sandbox error: {exc}")
            return _error(f"❌ Failed to create '{filename}': {exc}")

        try:
            _id, download_url, size_kb = await _store_document(
                user_id, session_id, filename, file_bytes, _XLSX_MIME
            )
        except Exception as exc:  # noqa: BLE001 - storage failure is terminal
            logger.error(f"create_excel_spreadsheet storage error: {exc}")
            return _error(f"❌ Created '{filename}' but failed to save it: {exc}")

        return _download_card(filename, download_url, size_kb, "Created")

    return create_excel_spreadsheet


def make_modify_excel_spreadsheet_tool(session_id: str, user_id: str):
    """Create a ``modify_excel_spreadsheet`` tool bound to the given identity."""

    @tool
    async def modify_excel_spreadsheet(
        spreadsheet_name: str,
        python_code: str,
        output_name: Optional[str] = None,
    ) -> Any:
        """Modify an existing Excel (.xlsx) spreadsheet with openpyxl code.

        Loads a workbook previously created in this chat, runs your openpyxl
        code against it, and saves the result (as a new file so the original is
        preserved). Returns a download card.

        Use ``list_excel_spreadsheets`` first if you are unsure of the exact
        name.

        Args:
            spreadsheet_name: Name of the existing workbook to edit (with or
                without the .xlsx extension), e.g. "sales-2026".
            python_code: openpyxl code that edits the workbook. The loaded
                workbook is available as ``wb = load_workbook(...)`` and its
                active sheet as ``ws = wb.active`` — do NOT call
                ``load_workbook()`` or ``wb.save()`` yourself. Access other
                sheets with ``wb['SheetName']`` and add sheets with
                ``wb.create_sheet('Name')``. ``Font``, ``PatternFill``,
                ``Alignment``, ``Border``, ``Side``, ``get_column_letter`` and
                the chart classes are already imported.

                Example (append rows to the active sheet):
                    ws.append(['Q3', 140])
                    ws.append(['Q4', 160])

            output_name: Optional name (without extension) for the edited copy.
                Defaults to the source name (a new versioned copy is saved).

        Returns:
            An inline download card for the edited workbook.
        """
        code_interpreter_id = _get_code_interpreter_id()
        if not code_interpreter_id:
            return _error(_NO_CI_MESSAGE)
        if not _storage_configured():
            return _error(_NO_STORAGE_MESSAGE)

        source = await _find_excel_spreadsheet(user_id, session_id, spreadsheet_name)
        if source is None:
            return _error(
                f"❌ No Excel spreadsheet named '{spreadsheet_name}' was found in "
                "this chat. Use list_excel_spreadsheets to see what's available."
            )

        out_base = output_name or source.filename
        if out_base.lower().endswith(".xlsx"):
            out_base = out_base[: -len(".xlsx")]
        is_valid, error_msg = _validate_document_name(out_base)
        if not is_valid:
            return _error(
                f"❌ Invalid output name '{out_base}': {error_msg}"
            )
        output_filename = f"{out_base}.xlsx"

        try:
            source_bytes = await asyncio.to_thread(
                _download_s3_bytes, source.s3_bucket, source.s3_key
            )
            file_bytes = await asyncio.to_thread(
                _modify_xlsx_bytes,
                code_interpreter_id,
                source_bytes,
                python_code,
                output_filename,
            )
        except _DocGenError as exc:
            return _error(
                f"❌ Failed to modify '{source.filename}'.\n\n```\n{exc}\n```\n\n"
                "Check the openpyxl code for errors."
            )
        except Exception as exc:  # noqa: BLE001 - surface any sandbox error
            logger.error(f"modify_excel_spreadsheet error: {exc}")
            return _error(f"❌ Failed to modify '{source.filename}': {exc}")

        try:
            _id, download_url, size_kb = await _store_document(
                user_id, session_id, output_filename, file_bytes, _XLSX_MIME
            )
        except Exception as exc:  # noqa: BLE001 - storage failure is terminal
            logger.error(f"modify_excel_spreadsheet storage error: {exc}")
            return _error(
                f"❌ Modified '{source.filename}' but failed to save it: {exc}"
            )

        return _download_card(output_filename, download_url, size_kb, "Updated")

    return modify_excel_spreadsheet


def make_list_excel_spreadsheets_tool(session_id: str, user_id: str):
    """Create a ``list_excel_spreadsheets`` tool bound to the given identity."""

    @tool
    async def list_excel_spreadsheets() -> Dict[str, Any]:
        """List the Excel (.xlsx) spreadsheets available in this chat.

        Returns the file names and sizes of workbooks created or modified in
        this conversation. Use the names with modify_excel_spreadsheet or
        read_excel_spreadsheet.
        """
        from apis.shared.files import FileStatus, get_file_upload_repository

        files = await get_file_upload_repository().list_session_files(
            session_id, status=FileStatus.READY
        )
        seen: set[str] = set()
        rows = []
        for meta in files:  # newest-first
            if meta.user_id != user_id or meta.mime_type != _XLSX_MIME:
                continue
            if meta.filename in seen:
                continue
            seen.add(meta.filename)
            rows.append(f"- {meta.filename} ({meta.size_bytes / 1024:.1f} KB)")

        if not rows:
            text = (
                "No Excel spreadsheets in this chat yet. Use "
                "create_excel_spreadsheet to make one."
            )
        else:
            text = "Excel spreadsheets in this chat:\n" + "\n".join(rows)
        return {"content": [{"text": text}], "status": "success"}

    return list_excel_spreadsheets


def make_read_excel_spreadsheet_tool(session_id: str, user_id: str):
    """Create a ``read_excel_spreadsheet`` tool bound to the given identity."""

    @tool
    async def read_excel_spreadsheet(spreadsheet_name: str) -> Dict[str, Any]:
        """Read the cell values of an existing Excel (.xlsx) spreadsheet.

        Extracts each sheet's rows (pipe-delimited, cached formula values when
        present) from a workbook created in this chat so you can reference or
        summarize its contents. Use list_excel_spreadsheets first if unsure of
        the exact name.

        Args:
            spreadsheet_name: Name of the workbook to read (with or without the
                .xlsx extension), e.g. "sales-2026".

        Returns:
            The workbook's cell values, grouped by sheet.
        """
        code_interpreter_id = _get_code_interpreter_id()
        if not code_interpreter_id:
            return _error(_NO_CI_MESSAGE)
        if not _storage_configured():
            return _error(_NO_STORAGE_MESSAGE)

        source = await _find_excel_spreadsheet(user_id, session_id, spreadsheet_name)
        if source is None:
            return _error(
                f"❌ No Excel spreadsheet named '{spreadsheet_name}' was found in "
                "this chat. Use list_excel_spreadsheets to see what's available."
            )

        try:
            source_bytes = await asyncio.to_thread(
                _download_s3_bytes, source.s3_bucket, source.s3_key
            )
            text = await asyncio.to_thread(
                _extract_xlsx_text, code_interpreter_id, source_bytes
            )
        except _DocGenError as exc:
            return _error(f"❌ Failed to read '{source.filename}': {exc}")
        except Exception as exc:  # noqa: BLE001 - surface any sandbox error
            logger.error(f"read_excel_spreadsheet error: {exc}")
            return _error(f"❌ Failed to read '{source.filename}': {exc}")

        body = text or "(The workbook has no extractable cell values.)"
        return {
            "content": [
                {"text": f"Content of {source.filename}:\n\n{body}"}
            ],
            "status": "success",
        }

    return read_excel_spreadsheet
